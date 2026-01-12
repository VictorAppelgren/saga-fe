#!/usr/bin/env python3
"""
Extract financial data from saga_budget.xlsx and generate TypeScript file.
Run this script before building to keep FinancialData.ts in sync with Excel.

Usage: python scripts/extract-financial-data.py
"""

import json
import os

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    os.system("pip install openpyxl")
    import openpyxl

EXCEL_PATH = "pitch-files/saga_budget.xlsx"
OUTPUT_PATH = "src/lib/components/pitch/FinancialData.ts"

def main():
    print(f"Reading {EXCEL_PATH}...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=False)

    # Get assumptions
    ws_assumptions = wb['Assumptions']
    RAISE_AMOUNT = ws_assumptions.cell(5, 2).value  # 35,000,000 SEK
    DILUTION = ws_assumptions.cell(6, 2).value  # 0.20
    AVG_SALARY = ws_assumptions.cell(11, 2).value  # 90,000 SEK/month
    ARBETSGIVARAVGIFT = ws_assumptions.cell(12, 2).value  # 0.32
    MONTHLY_LICENSE = ws_assumptions.cell(16, 2).value  # 100,000 SEK

    # Operating costs
    COMPUTE_BASE = ws_assumptions.cell(20, 2).value
    COMPUTE_PER_CUST = ws_assumptions.cell(21, 2).value
    DATA_BASE = ws_assumptions.cell(22, 2).value
    DATA_PER_CUST = ws_assumptions.cell(23, 2).value
    INFRA_BASE = ws_assumptions.cell(24, 2).value
    INFRA_PER_EMP = ws_assumptions.cell(25, 2).value
    OFFICE_PER_EMP = ws_assumptions.cell(26, 2).value
    SALES_BASE = ws_assumptions.cell(27, 2).value
    SALES_PER_CUST = ws_assumptions.cell(28, 2).value
    ADMIN_BASE = ws_assumptions.cell(29, 2).value
    BUFFER_PCT = ws_assumptions.cell(30, 2).value

    # Get inputs
    ws_inputs = wb['Inputs']
    new_customers = []
    team_size = []
    for col in range(2, 26):  # B to Y (M1-M24)
        new_customers.append(int(ws_inputs.cell(4, col).value or 0))
        team_size.append(int(ws_inputs.cell(7, col).value or 0))

    # Calculate derived values
    COST_PER_EMPLOYEE = AVG_SALARY * (1 + ARBETSGIVARAVGIFT)
    PRE_MONEY = RAISE_AMOUNT / DILUTION - RAISE_AMOUNT
    POST_MONEY = PRE_MONEY + RAISE_AMOUNT

    # Calculate cumulative customers
    total_customers = []
    cumulative = 0
    for nc in new_customers:
        cumulative += nc
        total_customers.append(cumulative)

    # Calculate monthly revenue
    monthly_revenue = [tc * MONTHLY_LICENSE for tc in total_customers]
    arr_run_rate = [mr * 12 for mr in monthly_revenue]

    # Calculate expenses
    salaries = [ts * COST_PER_EMPLOYEE for ts in team_size]
    compute = [COMPUTE_BASE + tc * COMPUTE_PER_CUST for tc in total_customers]
    data = [DATA_BASE + tc * DATA_PER_CUST for tc in total_customers]
    infra = [INFRA_BASE + ts * INFRA_PER_EMP for ts in team_size]
    office = [ts * OFFICE_PER_EMP for ts in team_size]
    sales = [SALES_BASE + tc * SALES_PER_CUST for tc in total_customers]
    admin = [ADMIN_BASE] * 24

    subtotal_expenses = []
    for i in range(24):
        sub = salaries[i] + compute[i] + data[i] + infra[i] + office[i] + sales[i] + admin[i]
        subtotal_expenses.append(sub)

    buffer = [int(se * BUFFER_PCT) for se in subtotal_expenses]
    total_expenses = [int(se + b) for se, b in zip(subtotal_expenses, buffer)]

    # Calculate cashflow
    net_cashflow = [int(mr - te) for mr, te in zip(monthly_revenue, total_expenses)]

    cash_balance = []
    balance = RAISE_AMOUNT
    for ncf in net_cashflow:
        balance += ncf
        cash_balance.append(int(balance))

    # Find breakeven month (first positive net cashflow)
    breakeven = 24
    for i, ncf in enumerate(net_cashflow):
        if ncf > 0:
            breakeven = i + 1
            break

    # Generate TypeScript
    ts_content = f'''// AUTO-GENERATED from saga_budget.xlsx
// Run: python scripts/extract-financial-data.py
// All figures in SEK

export const RAISE_INFO = {{
  amount: {int(RAISE_AMOUNT)},
  preMoney: {int(PRE_MONEY)},
  postMoney: {int(POST_MONEY)},
  dilution: {DILUTION},
  runway: 24,
}};

export const CASHFLOW = {{
  months: {list(range(1, 25))},
  totalCustomers: {total_customers},
  teamSize: {team_size},
  monthlyRevenue: {[int(x) for x in monthly_revenue]},
  arrRunRate: {[int(x) for x in arr_run_rate]},
  totalExpenses: {total_expenses},
  netCashflow: {net_cashflow},
  cashBalance: {cash_balance},
}};

export const MILESTONES = {{
  y1Customers: {total_customers[11]},
  y1ARR: {int(arr_run_rate[11])},
  y1Team: {team_size[11]},
  y2Customers: {total_customers[23]},
  y2ARR: {int(arr_run_rate[23])},
  y2Team: {team_size[23]},
  breakeven: {breakeven},
}};

export function formatSEK(value: number, compact = false): string {{
  if (compact) {{
    if (value >= 1_000_000_000) return `${{(value / 1_000_000_000).toFixed(1)}}B`;
    if (value >= 1_000_000) return `${{(value / 1_000_000).toFixed(1)}}M`;
    if (value >= 1_000) return `${{(value / 1_000).toFixed(0)}}K`;
    return value.toFixed(0);
  }}
  return new Intl.NumberFormat('sv-SE', {{ maximumFractionDigits: 0 }}).format(value) + ' SEK';
}}

export function formatPercent(value: number): string {{
  return `${{(value * 100).toFixed(0)}}%`;
}}

// Legacy exports for compatibility
export const RAISE_SCENARIOS = [RAISE_INFO];
export const CASHFLOW_24M = CASHFLOW;
export function formatCurrency(value: number, compact = false): string {{
  return formatSEK(value, compact);
}}
'''

    print(f"Writing {OUTPUT_PATH}...")
    with open(OUTPUT_PATH, 'w') as f:
        f.write(ts_content)

    print(f"\nExtracted data from Excel:")
    print(f"  Raise: {RAISE_AMOUNT:,.0f} SEK")
    print(f"  Pre-Money: {PRE_MONEY:,.0f} SEK")
    print(f"  Y1 Customers: {total_customers[11]}")
    print(f"  Y2 Customers: {total_customers[23]}")
    print(f"  Y2 ARR: {arr_run_rate[23]:,.0f} SEK")
    print(f"  Breakeven: Month {breakeven}")
    print(f"\nDone! {OUTPUT_PATH} updated.")

if __name__ == "__main__":
    main()
